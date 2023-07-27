'''
인기 검색 종목 5 스크래핑
Excel에 기록
'''
import requests
from bs4 import BeautifulSoup

from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill


def naver_finance():
    url = 'https://finance.naver.com/'
    datas = []
    try:
        response = requests.get(url)
        print(f'response.status_code: {response.status_code}')  # response.status_code: 200

        if response.status_code == 200:  # 성공적인 요청
            html = response.text
            # print(f'html: {html}')

            # BeautifulSoup 객체 생성
            soup = BeautifulSoup(html, 'html.parser')

            tbody = soup.select_one('div.aside_popular > table.tbl_home > tbody')
            # print(f'tbody: {tbody}')

            # tr 모두 선택
            trs = tbody.select('tr')

            # 인기 검색 종목 5
            for tr in trs:
                # print(tr)
                name = tr.select_one('th > a').text  # 종목
                current_price = tr.select_one('td').get_text()  # 주가
                change_price = tr.select_one('td > span').text  # 등락금액
                change_direction = tr['class'][0]
                print(f'{name}: {current_price} ({change_price}{change_direction})')
                datas.append([name, current_price, change_direction, change_price])

            print('-' * 50)
            # ----------- 정민 코드 ------------
            names = soup.select('.aside_area.aside_popular .tbl_home th a')
            for name in names:
                pass
                # print(name.text)
        # ---------------------------------
        else:
            print('상태코드를 확인하세요. {0}'.format(response.status_code))

    except Exception as e:
        print('Exception', str(e))

    return datas


def write_excel(datas):
    try:
        # Workbook 생성
        workbook = Workbook()

        # Get active sheet
        # sheet = workbook.active
        sheet = workbook.create_sheet('PCWK_20230727')

        # header
        sheet['A1'] = '종목'
        sheet['B1'] = '현재가'
        sheet['C1'] = '등락'
        sheet['D1'] = '등락가'

        # 스타일 지정
        font = Font(name='Arial', bold=True, color='000000')
        align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
        align_right = Alignment(horizontal='right', vertical='center', wrap_text=True)
        border = Border(
            left=Side(border_style="double", color="000000"),
            right=Side(border_style="double", color="000000"),
            top=Side(border_style="double", color="000000"),
            bottom=Side(border_style="double", color="000000")
        )

        sheet['A1'].font = font
        sheet['B1'].font = font
        sheet['C1'].font = font
        sheet['D1'].font = font

        sheet['A1'].alignment = align_center
        sheet['B1'].alignment = align_center
        sheet['C1'].alignment = align_center
        sheet['D1'].alignment = align_center

        sheet['A1'].border = border
        sheet['B1'].border = border
        sheet['C1'].border = border
        sheet['D1'].border = border

        for data in datas:
            sheet.append(data)

        workbook.save('naver_finance_0729.xlsx')

    except Exception as e:
        print('Exception', str(e))

    print('naver_finance_0729.xlsx 생성')


def main():
    datas = naver_finance()
    print(f'datas: {datas}')
    write_excel(datas)


main()
