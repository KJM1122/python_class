'''
엑셀에 입력
'''
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill


def style_write_excel():
    # Workbook 생성
    workbook = Workbook()

    # Get active sheet
    # sheet = workbook.active  # 첫 번째 sheet에 생성
    sheet = workbook.create_sheet('PCWK_20230727')  # 두 번째 sheet에 이름 바뀌며 생성

    sheet['A1'] = 'Hello'
    sheet['B1'] = 'World'

    sheet['A2'] = '100'
    sheet['B2'] = '200'

    # style -----------------------
    font = Font(bold=True)
    sheet['A1'].font = font

    font = Font(italic=True)
    sheet['B1'].font = font

    border = Border(
        left=Side(border_style="double", color="000000"),
        right=Side(border_style="double", color="000000"),
        top=Side(border_style="double", color="000000"),
        bottom=Side(border_style="double", color="000000")
    )
    sheet['B2'].border = border

    alignment = Alignment(horizontal="center", vertical="center")
    sheet['A1'].alignment = alignment

    fill = PatternFill(start_color="0000ff", end_color="0000ff", fill_type="solid")
    sheet['A2'].fill = fill
    # style -----------------------

    workbook.save('pcwk_excel_0729.xlsx')


def main():
    style_write_excel()


main()
